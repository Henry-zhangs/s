import os
import cv2
import torch
import streamlit as st
from PIL import Image
from openpyxl import Workbook
from openpyxl.styles import Font
import datetime
import torchvision.models as m
from torchvision import transforms
import tempfile


# Constants
LABELS = {
    "0": "Convusion",
    "1": "anxiety",
    "2": "body_twitching",
    "3": "exploratory_moving",
    "4": "extend_limbs",
    "5": "head_shaking",
    "6": "moderate_dysphea",
    "7": "scratching",
    "8": "severe_dysphea",
    "9": "washing_face"
}


def create_model(model, num, weights):
    net = m.efficientnet_b0(weights=m.EfficientNet_B0_Weights.DEFAULT if weights else False, progress=True)
    tmp = list(net.classifier)[-1].in_features
    net.classifier = torch.nn.Linear(tmp, num, bias=True)
    return net


def data_trans(train_mean=[0.485, 0.456, 0.406], train_std=[0.229, 0.224, 0.225]):
    train_transform = transforms.Compose([
        transforms.Resize(256),
        transforms.RandomRotation(90),
        transforms.CenterCrop(224),
        transforms.ToTensor(),
        transforms.Normalize(train_mean, train_std)
    ])

    test_transform = transforms.Compose([
        transforms.Resize(256),
        transforms.CenterCrop(224),
        transforms.ToTensor(),
        transforms.Normalize(train_mean, train_std)
    ])

    return train_transform, test_transform


def get_device():
    device = torch.device("cuda" if torch.cuda.is_available() else "cpu")
    st.sidebar.info(f"Using device: {device}")
    return device


def process_video(video_file, save_video, save_excel, progress_bar, status_text):
    device = get_device()
    _, data_transform = data_trans()

    # Create model
    net = create_model(model='b0', num=len(LABELS), weights=False)
    net.load_state_dict(torch.load('./runs/weights/best.pth'), strict=False)
    net.to(device)
    net.eval()

    # Create temp file for video
    temp_video = tempfile.NamedTemporaryFile(delete=False, suffix='.mp4')
    temp_video.write(video_file.read())
    temp_video.close()

    # Video input
    cap = cv2.VideoCapture(temp_video.name)
    if not cap.isOpened():
        status_text.error("无法打开视频文件")
        return None, None

    # Get video info
    fps = cap.get(cv2.CAP_PROP_FPS)
    width = int(cap.get(cv2.CAP_PROP_FRAME_WIDTH))
    height = int(cap.get(cv2.CAP_PROP_FRAME_HEIGHT))
    total_frames = int(cap.get(cv2.CAP_PROP_FRAME_COUNT))

    # Create Excel workbook if needed
    if save_excel:
        wb = Workbook()
        ws = wb.active
        ws.title = "Video Classification Results"
        headers = ["时间范围(秒)", "识别类别"]
        for col, header in enumerate(headers, 1):
            ws.cell(row=1, column=col, value=header).font = Font(bold=True)

    # Create video output if needed
    output_video_path = None
    if save_video:
        timestamp = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
        output_video_path = os.path.splitext(video_file.name)[0] + f"_output_{timestamp}.mp4"
        fourcc = cv2.VideoWriter_fourcc(*'mp4v')
        out = cv2.VideoWriter(output_video_path, fourcc, fps, (width, height))

    current_second = -1
    row_idx = 2
    last_class = None
    start_time = 0
    end_time = 0

    frame_count = 0
    video_placeholder = st.empty()
    stop_button = st.sidebar.button("停止处理")

    with torch.no_grad():
        while cap.isOpened():
            if stop_button:
                status_text.warning("处理已停止")
                break

            ret, frame = cap.read()
            if not ret:
                if last_class is not None and save_excel:
                    time_range = f"{start_time}~{end_time}" if start_time != end_time else str(start_time)
                    ws.cell(row=row_idx, column=1, value=time_range)
                    ws.cell(row=row_idx, column=2, value=last_class)
                break

            frame_count += 1
            progress = int((frame_count / total_frames) * 100)
            progress_bar.progress(progress)
            status_text.text(f"处理中... {progress}% 完成")

            current_time = frame_count / fps
            second = int(current_time)

            # Convert frame to PIL Image and preprocess
            pil_img = Image.fromarray(cv2.cvtColor(frame, cv2.COLOR_BGR2RGB))
            img = data_transform(pil_img)
            img = torch.unsqueeze(img, dim=0)

            # Predict
            output = net(img.to(device))
            output = torch.softmax(output, dim=1)
            p, index = torch.topk(output, k=3)
            current_class = LABELS[str(index.to("cpu").numpy()[0][0])]

            # Display top 3 probabilities in the frame
            text_y = 30
            for i in range(3):
                class_idx = index.to("cpu").numpy()[0][i]
                class_name = LABELS[str(class_idx)]
                prob = p.to("cpu").numpy()[0][i]
                text = f'{class_name}: {prob:.4f}'
                cv2.putText(frame, text, (10, text_y), cv2.FONT_HERSHEY_SIMPLEX, 0.8, (0, 255, 0), 2)
                text_y += 30

            # Check results every second
            if second != current_second:
                current_second = second

                if current_class != last_class:
                    if last_class is not None and save_excel:
                        time_range = f"{start_time}~{end_time}" if start_time != end_time else str(start_time)
                        ws.cell(row=row_idx, column=1, value=time_range)
                        ws.cell(row=row_idx, column=2, value=last_class)
                        row_idx += 1
                    start_time = second
                    last_class = current_class
                end_time = second

            # Display the frame
            rgb_image = cv2.cvtColor(frame, cv2.COLOR_BGR2RGB)
            video_placeholder.image(rgb_image, channels="RGB", use_container_width=True)

            # Save result if needed
            if save_video:
                out.write(frame)

    # Save Excel file
    output_excel_path = None
    if save_excel:
        timestamp = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
        output_excel_path = os.path.splitext(video_file.name)[0] + f"_classification_{timestamp}.xlsx"
        wb.save(output_excel_path)

    # Release resources
    cap.release()
    if save_video:
        out.release()

    # Clean up temp file
    os.unlink(temp_video.name)

    return output_video_path, output_excel_path


def main():
    st.set_page_config(
        page_title="智能视频分类工具",
        page_icon="🎬",
        layout="wide",
        initial_sidebar_state="expanded"
    )

    st.title("智能视频分类工具")

    # Sidebar for controls
    with st.sidebar:
        st.header("视频文件选择")
        video_file = st.file_uploader(
            "选择视频文件",
            type=["mp4", "avi", "mov"],
            label_visibility="collapsed"
        )

        st.header("输出选项")
        save_video = st.checkbox("保存识别结果视频", value=True)
        save_excel = st.checkbox("保存Excel分类结果", value=True)

        if st.button("开始分类", use_container_width=True):
            if not video_file:
                st.error("请先选择视频文件")
            else:
                st.session_state.processing = True

    # Main content area
    if 'processing' in st.session_state and st.session_state.processing:
        progress_bar = st.progress(0)
        status_text = st.empty()

        try:
            output_video, output_excel = process_video(
                video_file,
                save_video,
                save_excel,
                progress_bar,
                status_text
            )

            message = "处理完成！"
            if output_video:
                message += f"\n视频已保存到: {output_video}"
            if output_excel:
                message += f"\nExcel已保存到: {output_excel}"

            status_text.success(message)
            st.balloons()

        except Exception as e:
            status_text.error(f"处理过程中发生错误: {str(e)}")

        finally:
            st.session_state.processing = False

    # Display video info if selected
    if video_file and 'processing' not in st.session_state:
        st.subheader("视频信息")
        st.text(f"文件名: {video_file.name}")
        st.text(f"文件大小: {video_file.size / (1024 * 1024):.2f} MB")

        # Display first frame as preview
        temp_video = tempfile.NamedTemporaryFile(delete=False, suffix='.mp4')
        temp_video.write(video_file.read())
        temp_video.close()

        cap = cv2.VideoCapture(temp_video.name)
        if cap.isOpened():
            ret, frame = cap.read()
            if ret:
                rgb_image = cv2.cvtColor(frame, cv2.COLOR_BGR2RGB)
                st.image(rgb_image, caption="视频预览 (第一帧)", channels="RGB",use_container_width=True)

        cap.release()
        os.unlink(temp_video.name)

        # Reset file pointer
        video_file.seek(0)


if __name__ == '__main__':
    main()
